import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.drawing.image import Image

class MathsPreReqs:
    def __init__(self):
        pass
    #def __init__(self, filepath, SheetName):
        #self.filepath = filepath
        #self.SheetName = SheetName
        #self.wb = load_workbook(filepath)
        #self.ws = self.wb[SheetName]

    def GetPercent(self, PercentInput: float):
        Percent = PercentInput / 100
        return round(Percent,2)

    def GetDiscount(self, PricingType: str, Percent: float, VendorLineDiscount, UnitCost, UnitList):
        PricingType = PricingType.strip().upper()
        if UnitList == 0:
            LineDiscount = 0
        else:
            if PricingType == "HOLD BACK":
                LineDiscount = (VendorLineDiscount / 100) - Percent
            elif PricingType == "MARKUP":
                LineDiscount = 1 - ((UnitCost * (Percent + 1)) / UnitList)
            elif PricingType == "MARGIN":
                if Percent == 1:
                    print("⚠️ Percent  is 100!")
                    LineDiscount = 1
                else:
                    LineDiscount = 1 - ((UnitCost / (1 - Percent)) / UnitList)
            else:
                LineDiscount = 0
        
        if LineDiscount < 0 or LineDiscount > 1:
            print(f"⚠️ Discount out of bounds! Clamped to {LineDiscount:.2%}")
        LineDiscount = max(0, min(LineDiscount, 1)) # clamp to safe range to avoid negaive unit NPs
        print(f"📊 {PricingType} Discount → Vendor: {VendorLineDiscount}% | Calculated: {LineDiscount:.2%}")
        return round(LineDiscount,2)
        

class Maths:
    def __init__(self):
        pass

    def GetUnitNP(self, LineDiscount, UnitList):
        UnitNP = UnitList - (UnitList*LineDiscount)
        return round(UnitNP,2)
            
    def GetLineExtendedNP(self, UnitNP, QTY, LineDuration, PTerm):
        if PTerm == 0:
            print("⚠️ Pricing Term  is 0!")
            LineExtendedNP = 0
        else:
            LineExtendedNP = (UnitNP * QTY * LineDuration) / PTerm
            return round(LineExtendedNP,2)

    def GetUnitMonthlyNP(self, UnitNP, PTerm): #this will come in handy for mods and situations with credits
        if PTerm == 0:
            print("⚠️ Pricing Term  is 0!")
            UnitMonthlyNP = 0
        else:
            UnitMonthlyNP = UnitNP / PTerm
        return round(UnitMonthlyNP,2)

    def GetLineMonthlyNP(self, TotalDuration, LineExtendedNP):
            if TotalDuration == 0:
                print("⚠️ Duration is 0!")
                LineMonthlyNP = 0
            else:
                LineMonthlyNP = LineExtendedNP / TotalDuration
            return round(LineMonthlyNP,2)

    def GetPaymentQTY(self, BillingType: str, Duration: float): #this will be used new, renewals, and mods... Total number of payments over the entire duration depending on the Billing
        BillingType = BillingType.strip().title()
        if BillingType == "Prepaid":
            PaymentQTY = 1
        elif BillingType =="Monthly":
            PaymentQTY =  Duration
        elif BillingType == "Annual":
            PaymentQTY = Duration / 12
        elif BillingType == "Quarterly":
            PaymentQTY = (Duration / 12)*4
        else:
            PaymentQTY = 1
        return round(PaymentQTY, 0)
        
    def GetBillingAmount(self, ExtendedNP, PaymentQTY): #this will be used new, renewals, and mods
        if PaymentQTY == 0:
            print("⚠️ Number of payments is 0!")
            BillingAmount = ExtendedNP
        else:
            BillingAmount = ExtendedNP / PaymentQTY
        return round(BillingAmount,2)


#------ PRORATIONS

#    def GetProratedPrice(self, LineMonthlyNP, ExistingRemainingDuration, NewRemainingDuration):
#        ExistingLineProratedNP = LineMonthlyNP * ExistingRemainingDuration
#        NewLineProratedNP = LineMonthlyNP * NewRemainingDuration
#        return ExistingLineProratedNP, NewLineProratedNP
    
#    def GetLineCM(self, TFqty, ExistingLineProratedNP, TransactionType):
#        if TransactionType == "True Forward":
#            if TFqty == 0: 
#                LineCM = 0
#            else:
#                LineCM = ExistingLineProratedNP
#        elif TransactionType =="Modification":
#            LineCM = ExistingLineProratedNP
#        else:
#            LineCM = 0
#        return LineCM 

#    def GetLineINV(self, TFqty, NewLineProratedNP, TransactionType):
#        if TransactionType == "True Forward":
#            if TFqty == 0: 
#                LineINV = 0
#            else:
#                LineINV = NewLineProratedNP
#        elif TransactionType =="Modification":
##            LineINV = NewLineProratedNP
#       else:
#           LineINV = 0
# #       return LineINV 

#True Ups
#   def GetTFqty(self, vsRETURNED, ExistingQTY, NewQTY):
#       TFqty = 0 if vsRETURNED == 0 and ExistingQTY > NewQTY else NewQTY - ExistingQTY
#       return TFqty

#   def GetConsumptionStatus(self, TFqty, vsADD, vsRETURNED, TIER):
#       if TFqty == 0:
#           Consumption =  "Within Allocation"
#       elif TFqty > 0:
#           if vsADD > 0:
#               Consumption = "(+) License Shift"
#           elif vsADD == 0:
#               Consumption = "Allocation Exceeded"
#       elif TFqty < 0:
#           if TIER and TIER.upper() == "FIXED":
#               Consumption = "FIXED"
#           elif vsRETURNED > 0:
#               Consumption = "(-) License Shift"
#       else:
#           Consumption = "Unknown"
#       return Consumption


#    def GetLineDeltaNP(self, LineINV, LineCM): #for true ups
#        LineDeltaNP = LineINV - LineCM
#        return LineDeltaNP



#____ MODS
#    def GetExistingTotalNP(self):
#        return ExistingTotalNP

#    def GetNewTotalNP(self):
#        NewTotalNP = 
#        return NewTotalNP
    
#    def GetExistingDurationRemaining(self):
#        ExistingDurationRemaining = 
#        return ExistingDurationRemaining

#    def Get NewDurationRemaining(self):
#        NewDurationRemaining = 
#        return NewDurationRemaining

#    def Get ExistingToFrom(self): #string
#        ExistingToFrom = 
#        return ExistingToFrom

#    def GetNewToFrom(self):  #string
#        NewToFrom
#        return NewToFrom

#    def GetExistingBCNP(self):
#        ExistingBCNP = 
#        return ExistingBCNP

#    def GetNewBCNP(self):
#        NewBCNP = 
#        return NewBCNP


#    def GetEndDateExistingBC(self):
#    EndDateExistingBC
#    return NewTotalNP


#    def GetEndDateNewBC(self):
#    EndDateNewBC
#    return

#    def GetExistingRemainingBC(self):
#    ExistingRemainingBC
#    return

#    def GetNewRemainingBC(self):
#    NewRemainingBC
#    return

#   def GetExistingToFromBC(self):  #string
#   ExistingToFromBC
#   return

#    def GetNewToFromBC(self):  #string
#    NewToFromBC
#    return

#    def GetLineNetChange(self, LineINV, LineCM): #for modifications
#        LineNetChange = LineINV - LineCM
#        return LineNetChange
#------CREDDITS

#    def GetCreditsNP(self):
#    CreditsNP
#    return

#    def GetNoCreditsNP(self):
#    NoCreditsNP
#    return

#   def GetMonthlyCredit(self):
#   MonthlyCredit
#   return




#class ModMaths:
#def __init__(self, filepath, sheet_name):
#    self.filepath = filepath
#    self.sheet_name = sheet_name
#    self.wb = load_workbook(filepath)
#    self.ws = self.wb[sheet_name]
 
#def Maths(self):
#    self.wb.save(self.filepath)


#class DeltaMaths:
#   def __init__(self, filepath, sheet_name):
#        self.filepath = filepath
#        self.sheet_name = sheet_name
#       self.wb = load_workbook(filepath)
#       self.ws = self.wb[sheet_name]


#def Maths(self):
#       self.wb.save(self.filepath)