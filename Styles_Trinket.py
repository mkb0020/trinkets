import pandas as pd
import tkinter as tk
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

def Decimals(df, columns):
    for col in columns:
        if col in df.columns:
            df[col] = df[col].map(lambda x: f"{float(x):.2f}" if pd.notnull(x) else "")
    return df


class DrippyKit:
    # ---------------------------- NUMBER FORMATS ---------------------------- 
    StandardCurrencyCols = {"UNIT LIST PRICE", "UNIT NET PRICE", "EXTENDED NET PRICE"}
    TDSummCurrencyCols = {"UNIT LIST PRICE", "UNIT NET PRICE", "ESTIMATED CREDIT", "ESTIMATED INVOICE",	"TRUE DELTA NET COST"}
    TDDeetsCurrencyCols = {"UNIT LIST PRICE", "UNIT NET PRICE","EXISTNG NET PRICE", "NEW NET PRICE", "EXISTING PRORATED PRICE", "NEW PRORATED PRICE", "ESTIMATED CREDIT", "ESTIMATED INVOICE", "TRUE DELTA NET COST"}
    NewSummaryCurrencyCols = StandardCurrencyCols #placehlder
    NewDeetsCurrencyCols = StandardCurrencyCols #placehlder
    RenewalSummaryCurrencyCols = StandardCurrencyCols #placehlder
    RenewalDeetsCurrencyCols = StandardCurrencyCols #placehlder
    ModSummaryCurrencyCols = StandardCurrencyCols #placehlder
    ModDeetsCurrencyCols = StandardCurrencyCols #placehlder 
    # ---------------------------- FILLS ---------------------------- 
    HeaderColor = PatternFill(start_color="4B1395", end_color="4B1395", fill_type="solid")
    SubHeaderColor = PatternFill(start_color="CBCEDF", end_color="CBCEDF", fill_type="solid")
    BodyColor = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    #Optinal color if needed: 9A9FC0
    # ---------------------------- BORDERS ---------------------------- 
    Border_IttyBitty = Side(style="thin", color="000000")
    Border_SheThicc = Side(style="thick", color="000000")
    Border_DoubleBottom = Side(border_style="double", color="000000")
    # ---------------------------- FONTS ----------------------------        
    MainHeaderFont = Font(name="Aptos Display", bold=True, color="FFFFFF", size=14)
    HeaderFont = Font(name="Aptos Narrow", bold=True, color="FFFFFF", size=10)
    SubHeaderFont = Font(name="Aptos Narrow", bold=True, color="000000", size=10)
    BodyFont = Font(name="Aptos Narrow", bold=False, color="000000", size=10)
    BodyBoldFont = Font(name="Aptos Narrow", bold=True, color="000000", size=10)
    # ---------------------------- ALIGNMENTS ---------------------------- 
    Middle = Alignment(horizontal="center", vertical="center", wrap_text=True)
    Lefty = Alignment(horizontal="left", vertical="center", indent=1)
    Righty = Alignment(horizontal="right", vertical="center", indent=1)


    HeaderStyle = {
        "font": HeaderFont,
        "alignment": Middle,
        "fill": HeaderColor,
        "border": Border_IttyBitty
    }

    ItemStyle = {
        "font": BodyFont,
        "alignment": Middle,
        "fill": BodyColor,
        "border": Border_IttyBitty
    }

    SubHeaderSytle = {
        "font": BodyBoldFont,
        "alignment": Righty,
        "fill": SubHeaderColor,
        "border": Border_IttyBitty
    }

    @staticmethod
    def GetTheDrip(cell, font=None, alignment=None, fill=None, border=None):
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if fill:
            cell.fill = fill
        if border:
            cell.border = border

    @staticmethod
    def Thicc(ws, FirstRow, LastRow, FirstColumn, LastColumn, ThiccBorder=Border_SheThicc):
        for row in ws.iter_rows(min_row=FirstRow, max_row=LastRow, min_col=FirstColumn, max_col=LastColumn):
            for cell in row:
                borders = {k: cell.border.__dict__[k] for k in ("left", "right", "top", "bottom")}
                if cell.row == 1: borders["top"] = ThiccBorder
                if cell.row == LastRow: borders["bottom"] = ThiccBorder
                if cell.column == 1: borders["left"] = ThiccBorder
                if cell.column == LastColumn: borders["right"] = ThiccBorder
                cell.border = Border(**borders)


class DeetsDrip:
    def __init__(self, filepath, sheet_name):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.wb = load_workbook(filepath)
        self.ws = self.wb[sheet_name]  
#-------------DETAILS HEADERS-------------         
    def HeaderLewk(self, HeaderRow):    
        self.ws.row_dimensions[1].height = 45
        for cell in HeaderRow:
            DrippyKit.ApplyDrip(cell, **DrippyKit.HeaderStyle)
    #-------------DETAILS ITEMS SECTION-------------   
    def ItemsLewk(self, FirstItemsRow, LastRow, LastColumn):    
        for row in range(FirstItemsRow, LastRow):
            self.ws.row_dimensions[row].height = 15
            for col in range(1, LastColumn):
                cell = self.ws.cell(row, col)
                DrippyKit.ApplyDrip(cell, **DrippyKit.ItemStyle)
    #-------------THICK BORDER-------------     
    def ThiccBorder(self, LastRow, LastColumn):        
        DrippyKit.Thicc(self.ws, 1, LastRow, 1, LastColumn)

    def drip(self):
        LastColumn = self.ws.max_column
        HeaderRow = next(self.ws.iter_rows(min_row=1, max_row=1))
        FirstItemsRow = 2  # Just use the row number
        LastRow = self.ws.max_row + 1
        self.HeaderLewk(HeaderRow, LastColumn)
        self.ItemsLewk(FirstItemsRow, LastRow)
        self.ThiccOutline(LastRow, LastColumn)
        self.wb.save(self.filepath)


class SumaryDrip:
    def __init__(self, filepath, sheet_name):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.wb = load_workbook(filepath)
        self.ws = self.wb[sheet_name]
    #-------------TITLE HEADER-------------
    def MainHeaderLewk(self, MainHeaderRow, LastColumn): 
        for cell in MainHeaderRow:
            DrippyKit.GetTheDrip(
            cell,
            font=DrippyKit.MainHeaderFont,
            alignment=DrippyKit.Middle,
            fill=DrippyKit.HeaderColor,
            border=DrippyKit.Border_IttyBitty
        )   
        self.ws.row_dimensions[1].height = 22
        DrippyKit.Thicc(self.ws, 1, 1, 1, LastColumn) #Apply Thick utline around main title header
        
        self.ws.row_dimensions[2].height = 5 #Row between the main title header and the general info section
    #-------------GENERAL INFO SECTION-------------
    def GenInfoLewk(self, FirstGenInfoRow, LastGenInfoRow, HeaderRow): 
        for row in range(FirstGenInfoRow, LastGenInfoRow): 
            self.ws.row_dimensions[row].height = 12
            for col in range(1): #Column A in the General Info Section
                cell = self.ws.cell(row, col)
                DrippyKit.GetTheDrip(
                    cell,
                    font=DrippyKit.SubHeaderFont,
                    alignment=DrippyKit.Lefty,
                    fill=DrippyKit.SubHeaderColor,
                    border=DrippyKit.Border_IttyBitty
                )
            for col in range(2): #Column B in the General Info Section
                    cell = self.ws.cell(row, col)
                    DrippyKit.GetTheDrip(
                    cell,
                    font=DrippyKit.BodyFont,
                    alignment=DrippyKit.Righty,
                    fill=DrippyKit.BodyColor,
                    border=DrippyKit.Border_IttyBitty
                )
        DrippyKit.Thicc(self.ws, FirstGenInfoRow, LastGenInfoRow, 1, 2) #Apply Thick Outline to General Info Section

        self.ws.row_dimensions[HeaderRow - 1].height = 5 #Row between the General Info and the items section
    #-------------SUMARY ITEMS SECTION-------------
    def ItemsLewk(self, LastColumn, HeaderRow, FirstItemsRow, LastItemsRow, TotalsRow): #Items Section - including headers and totals
        for col in range(1, LastColumn + 1): #Header Row
            cell = self.ws.cell(row=HeaderRow, column=col)
            DrippyKit.ApplyDrip(cell, **DrippyKit.HeaderStyle)   
        self.ws.row_dimensions[HeaderRow].height = 30 
        for row in range(FirstItemsRow, LastItemsRow): #Items
            self.ws.row_dimensions[row].height = 15
            for col in range(1, LastColumn):
                cell = self.ws.cell(row, col)
                DrippyKit.ApplyDrip(cell, **DrippyKit.ItemStyle)
               
        for col in range(1, LastColumn + 1): #Style Totals Row   
            cell = self.ws.cell(row=TotalsRow, column=col) 
            DrippyKit.ApplyDrip(cell, **DrippyKit.SubHeaderSytle)      
        self.ws.row_dimensions[TotalsRow].height = 15

        DrippyKit.Thicc(self.ws, HeaderRow, TotalsRow, 1, LastColumn) #Thick outline around items section - including the headers and totals row

        self.ws.row_dimensions[TotalsRow + 1].height = 5 #Row between the totals row and notes section
    #-------------NOTES SECTION-------------
    def NotesLewk(self, LastColumn, FirstNotesRow, LastNotesRow): #Notes Section    
        for col in range(1, LastColumn + 1):  #Style Notes Header
            cell = self.ws.cell(row=FirstNotesRow, column=col)
            DrippyKit.GetTheDrip(
            cell,
            font=DrippyKit.HeaderFont,
            alignment=DrippyKit.Lefty,
            fill=DrippyKit.HeaderColor,
            border=DrippyKit.Border_IttyBitty
            )
        self.ws.row_dimensions[FirstNotesRow].height = 15
        for row in range(FirstNotesRow+1, LastNotesRow): #Style Notes Section
            self.ws.row_dimensions[row].height = 15
            for col in range(1, LastColumn):
                cell = self.ws.cell(row, col)
                DrippyKit.GetTheDrip(
                cell,
                font=DrippyKit.BodyFont,
                alignment=DrippyKit.Lefty,
                fill=DrippyKit.SubHeaderColor,
                border=DrippyKit.Border_IttyBitty
            )
        DrippyKit.Thicc(self.ws, FirstNotesRow, LastNotesRow, 1, LastColumn) #Thick outline around notes section
    #-------------THICK BORDER-------------     
    def ThiccOutline(self, LastNotesRow, LastColumn): #Thick outline around The whole thing
        DrippyKit.Thicc(self.ws, 1, LastNotesRow, 1, LastColumn) 

    def drip(self):
        LastColumn = self.ws.max_column
        MainHeaderRow = next(self.ws.iter_rows(min_row=1, max_row=1))
        FirstGenInfoRow = 3
        LastGenInfoRow = 19 #this will need to be dynamic but we'll worry about that later!
        HeaderRow = LastGenInfoRow + 2
        FirstItemsRow = HeaderRow + 1
        LastItemsRow = self.ws.max_row - 5
        TotalsRow = LastItemsRow + 1
        FirstNotesRow = TotalsRow + 2
        LastNotesRow = FirstNotesRow + 3
        self.MainHeaderLewk(MainHeaderRow, LastColumn)
        self.GenInfoLewk(FirstGenInfoRow, LastGenInfoRow, HeaderRow)
        self.ItemsLewk(LastColumn, HeaderRow, FirstItemsRow, LastItemsRow, TotalsRow)
        self.NotesLewk(LastColumn, FirstNotesRow, LastNotesRow)
        self.ThiccOutline(LastNotesRow, LastColumn)
        self.wb.save(self.filepath)


    


